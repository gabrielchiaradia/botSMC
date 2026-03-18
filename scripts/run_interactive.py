#!/usr/bin/env python3
"""
scripts/run_interactive.py
──────────────────────────
Backtest interactivo: pregunta todas las opciones y corre
combinaciones múltiples automáticamente.

Uso:
    python scripts/run_interactive.py
"""

import subprocess
import sys
import os
import json
from pathlib import Path
from datetime import datetime
from itertools import product

# Colores
GREEN  = "\033[32m"
YELLOW = "\033[33m"
CYAN   = "\033[36m"
BOLD   = "\033[1m"
DIM    = "\033[2m"
RESET  = "\033[0m"

# Valores por defecto y opciones
SYMBOLS_DEFAULT   = ["ETHUSDT", "BTCUSDT", "SOLUSDT"]
TIMEFRAMES_ALL    = ["5m", "15m", "1h", "4h"]
STRATEGIES_ALL    = ["base", "ob_bos", "ema_filter", "ob_bos_ema"]
TRAILING_ALL      = ["none", "escalones", "atr", "hibrido"]


def ask(prompt, default="", options=None, multi=False, tipo=str):
    """Pregunta interactiva con opciones y soporte multi-valor."""
    hint = ""
    if options:
        hint = f" {DIM}[{', '.join(options)}]{RESET}"
    if default:
        hint += f" {DIM}(default: {default}){RESET}"
    if multi:
        hint += f" {DIM}(separar con coma para múltiples){RESET}"

    raw = input(f"\n{CYAN}▸ {prompt}{hint}\n  > {RESET}").strip()

    if not raw:
        raw = str(default)

    if multi:
        values = [v.strip() for v in raw.split(",") if v.strip()]
        if not values:
            values = [str(default)] if default else []
        # Validar contra opciones
        if options:
            valid = []
            for v in values:
                if v in options:
                    valid.append(v)
                else:
                    print(f"  {YELLOW}⚠ '{v}' no es válido, ignorado{RESET}")
            values = valid if valid else [str(default)]
        return [tipo(v) for v in values]
    else:
        if options and raw not in options:
            print(f"  {YELLOW}⚠ '{raw}' no es válido, usando default: {default}{RESET}")
            raw = str(default)
        return tipo(raw) if raw else tipo(default)


def ask_yn(prompt, default="s"):
    """Pregunta sí/no."""
    hint = " (S/n)" if default == "s" else " (s/N)"
    raw = input(f"\n{CYAN}▸ {prompt}{hint}\n  > {RESET}").strip().lower()
    if not raw:
        return default == "s"
    return raw in ("s", "si", "sí", "y", "yes")


def build_command(symbol, timeframe, dias, capital, strategy, trailing,
                  max_trades, windows, rr, score, ob_max_age=None):
    """Construye el comando de backtest."""
    cmd = [
        sys.executable, "scripts/run_backtest.py",
        "--symbol", symbol,
        "--timeframe", timeframe,
        "--dias", str(dias),
        "--capital", str(capital),
        "--strategy", strategy,
        "--trailing", trailing,
        "--max-trades", str(max_trades),
    ]
    if windows is not None:
        cmd.extend(["--windows", windows])
    if rr is not None:
        cmd.extend(["--rr", str(rr)])
    if score is not None:
        cmd.extend(["--score", str(score)])
    if ob_max_age is not None:
        cmd.extend(["--ob-max-age", str(ob_max_age)])
    return cmd


def format_combo(i, total, symbol, tf, strategy, trailing, rr, windows):
    """Formatea la descripción de una combinación."""
    parts = [f"{symbol}/{tf}", strategy, f"trail:{trailing}"]
    if rr:
        parts.append(f"RR 1:{rr}")
    if windows is not None:
        parts.append(f"w:{windows}" if windows else "24h")
    return f"[{i}/{total}] {' · '.join(parts)}"


def exportar_excel(archivos, excel_path):
    """Genera un Excel con los resultados de todos los backtests."""
    import re
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("Instalá openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtests"

    # Headers
    headers = ["Symbol", "TF", "Strategy", "Trailing", "Windows", "Días",
               "RR", "Retorno %", "Trades", "Win Rate %", "Profit Factor",
               "Max DD %", "Sharpe", "Avg Win", "Avg Loss", "Avg RR",
               "Cap Inicial", "Cap Final", "Archivo"]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    thin_border = Border(
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Datos
    green_font = Font(color="16A34A")
    red_font = Font(color="DC2626")

    def extraer_windows(nombre):
        m = re.search(r'_w(\d+-\d+(?:_\d+-\d+)*)', nombre)
        return m.group(1).replace('_', ',') if m else "24h"

    row = 2
    for path in archivos:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue

        if data.get("comparacion"):
            continue  # saltar systematic

        cap_ini = data.get("capital_inicial", 0)
        cap_fin = data.get("capital_final", 0)
        ret = (cap_fin - cap_ini) / cap_ini * 100 if cap_ini else 0
        nombre = Path(path).stem

        valores = [
            data.get("symbol", "?"),
            data.get("timeframe", "?"),
            data.get("perfil", "?"),
            data.get("trailing_mode", "none"),
            extraer_windows(nombre),
            data.get("dias", "?"),
            data.get("tp_rr_ratio", "?"),
            round(ret, 2),
            data.get("total_trades", 0),
            data.get("win_rate", 0),
            data.get("profit_factor", 0),
            data.get("max_drawdown_pct", 0),
            data.get("sharpe_ratio", 0),
            data.get("avg_win_usd", 0),
            data.get("avg_loss_usd", 0),
            data.get("avg_rr_real", 0),
            cap_ini,
            round(cap_fin, 2),
            Path(path).name,
        ]

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            # Colorear retorno
            if col == 8:
                cell.font = green_font if val >= 0 else red_font
                cell.number_format = '+0.00%;-0.00%'
            # Colorear PF
            if col == 11:
                cell.font = green_font if val >= 1.3 else red_font

        row += 1

    # Auto-ajustar anchos
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    # Filtros automáticos
    ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{row-1}"

    wb.save(excel_path)


def main():
    print(f"\n{'='*60}")
    print(f"{BOLD}{'  SMC BACKTEST INTERACTIVO':^60}{RESET}")
    print(f"{'='*60}")
    print(f"{DIM}  Configurá las opciones. Usá coma para múltiples valores.{RESET}")
    print(f"{DIM}  Enter = valor por defecto.{RESET}")

    # ── Preguntas ──────────────────────────────────────────
    symbols    = ask("Symbols", default="ETHUSDT",
                     options=SYMBOLS_DEFAULT + ["BNBUSDT", "XRPUSDT", "ADAUSDT", "DOGEUSDT"],
                     multi=True)

    timeframes = ask("Timeframes", default="15m",
                     options=TIMEFRAMES_ALL, multi=True)

    dias_list  = ask("Días de historia", default="360",
                     multi=True, tipo=int)

    capital    = ask("Capital inicial USDT", default="5000", tipo=float)

    strategies = ask("Estrategias", default="base",
                     options=STRATEGIES_ALL, multi=True)

    trailings  = ask("Trailing modes", default="none",
                     options=TRAILING_ALL, multi=True)

    max_trades_list = ask("Max trades simultáneos", default="1",
                          multi=True, tipo=int)

    # Windows
    use_windows = ask_yn("¿Definir ventanas horarias? (si no, usa .env)")
    windows_list = [None]
    if use_windows:
        raw = input(f"\n{CYAN}▸ Ventanas UTC (ej: 13-16). Vacío o 24h = sin filtro. Separar con ; para múltiples\n  > {RESET}").strip()
        if raw:
            windows_list = []
            for w in raw.split(";"):
                w = w.strip().strip('"').strip("'")  # quitar comillas
                if w.lower() in ("24h", "24hs", "todo", "all", ""):
                    windows_list.append("")
                else:
                    windows_list.append(w)
            if not windows_list:
                windows_list = [""]
        else:
            windows_list = [""]  # vacío = 24h sin filtro

    # RR
    use_rr = ask_yn("¿Definir RR ratio? (si no, usa .env)")
    rr_list = [None]
    if use_rr:
        rr_list = ask("RR ratios", default="1.5",
                       multi=True, tipo=float)

    # Score
    use_score = ask_yn("¿Definir score mínimo? (si no, usa .env)", default="n")
    score_list = [None]
    if use_score:
        score_list = ask("Score mínimo", default="50",
                          multi=True, tipo=int)

    # OB Max Age (solo aplica a ob_bos)
    ob_age_list = [None]
    if any(s in ("ob_bos", "ob_bos_ema") for s in strategies):
        use_age = ask_yn("¿Definir edad máxima del OB en velas? (solo ob_bos)", default="n")
        if use_age:
            ob_age_list = ask("OB max age (velas)", default="25",
                               multi=True, tipo=int)

    # ── Calcular combinaciones ─────────────────────────────
    combos = list(product(
        symbols, timeframes, dias_list, strategies, trailings,
        max_trades_list, windows_list, rr_list, score_list, ob_age_list
    ))

    print(f"\n{'─'*60}")
    print(f"{BOLD}  RESUMEN{RESET}")
    print(f"{'─'*60}")
    print(f"  Symbols:     {', '.join(symbols)}")
    print(f"  Timeframes:  {', '.join(timeframes)}")
    print(f"  Días:        {', '.join(str(d) for d in dias_list)}")
    print(f"  Capital:     ${capital:,.0f}")
    print(f"  Estrategias: {', '.join(strategies)}")
    print(f"  Trailing:    {', '.join(trailings)}")
    print(f"  Max trades:  {', '.join(str(m) for m in max_trades_list)}")
    print(f"  Windows:     {', '.join(str(w) for w in windows_list)}")
    print(f"  RR:          {', '.join(str(r) for r in rr_list)}")
    print(f"  Score:       {', '.join(str(s) for s in score_list)}")
    print(f"\n  {BOLD}{GREEN}Total combinaciones: {len(combos)}{RESET}")
    print(f"{'─'*60}")

    if not ask_yn("¿Ejecutar?"):
        print("  Cancelado.\n")
        return

    # ── Ejecutar ───────────────────────────────────────────
    resultados = []
    errores = []
    results_dir = Path(__file__).resolve().parents[1] / "backtest" / "results"

    # Snapshot de archivos antes de empezar
    archivos_antes = set(results_dir.glob("*.json")) if results_dir.exists() else set()

    for i, (sym, tf, dias, strat, trail, mt, win, rr, sc, age) in enumerate(combos, 1):
        desc = format_combo(i, len(combos), sym, tf, strat, trail, rr, win)
        print(f"\n{'═'*60}")
        print(f"  {GREEN}{desc}{RESET}")
        print(f"{'═'*60}")

        cmd = build_command(sym, tf, dias, capital, strat, trail, mt, win, rr, sc, age)

        try:
            result = subprocess.run(
                cmd,
                cwd=str(Path(__file__).resolve().parents[1]),
                timeout=600,  # 10 min max por backtest
            )
            if result.returncode == 0:
                resultados.append(desc)
            else:
                errores.append((desc, f"Exit code {result.returncode}"))
        except subprocess.TimeoutExpired:
            errores.append((desc, "Timeout (>10 min)"))
            print(f"  {YELLOW}⚠ Timeout, saltando...{RESET}")
        except Exception as e:
            errores.append((desc, str(e)))
            print(f"  {YELLOW}⚠ Error: {e}{RESET}")

    # ── Detectar archivos nuevos ───────────────────────────
    archivos_despues = set(results_dir.glob("*.json")) if results_dir.exists() else set()
    archivos_nuevos = sorted(archivos_despues - archivos_antes)

    # ── Guardar manifest con lista de archivos ─────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    manifest_path = results_dir / f"_run_{timestamp}.txt"
    if archivos_nuevos:
        try:
            with open(manifest_path, "w", encoding="utf-8") as f:
                f.write(f"# Interactive backtest run {timestamp}\n")
                f.write(f"# {len(archivos_nuevos)} resultados\n")
                for a in archivos_nuevos:
                    f.write(f"{a}\n")
            print(f"  {CYAN}Manifest: {manifest_path.name}{RESET}")
        except OSError:
            pass

    # ── Resumen final ──────────────────────────────────────
    print(f"\n\n{'═'*60}")
    print(f"{BOLD}{'  RESUMEN FINAL':^60}{RESET}")
    print(f"{'═'*60}")
    print(f"  {GREEN}✓ Completados: {len(resultados)}/{len(combos)}{RESET}")
    if errores:
        print(f"  {YELLOW}✗ Errores: {len(errores)}{RESET}")
        for desc, err in errores:
            print(f"    · {desc}: {err}")

    print(f"  Archivos generados: {len(archivos_nuevos)}")
    print(f"{'═'*60}")

    # ── Tabla comparativa ──────────────────────────────────
    if len(archivos_nuevos) >= 2:
        print(f"\n  {CYAN}Generando tabla comparativa...{RESET}\n")
        try:
            compare_cmd = [
                sys.executable, "scripts/compare_results.py",
                "--from-file", str(manifest_path),
                "--sort", "pf",
            ]
            subprocess.run(
                compare_cmd,
                cwd=str(Path(__file__).resolve().parents[1]),
            )
        except Exception as e:
            print(f"  {YELLOW}⚠ No se pudo generar comparación: {e}{RESET}")
    elif len(archivos_nuevos) == 1:
        print(f"\n  Resultado: {archivos_nuevos[0].name}")

    # ── Exportar Excel ─────────────────────────────────────
    if len(archivos_nuevos) >= 1:
        excel_path = results_dir / f"_run_{timestamp}.xlsx"
        try:
            exportar_excel(archivos_nuevos, excel_path)
            print(f"\n  {GREEN}Excel exportado: {excel_path.name}{RESET}")
        except Exception as e:
            print(f"  {YELLOW}⚠ No se pudo exportar Excel: {e}{RESET}")
            print(f"  {DIM}Tip: pip install openpyxl{RESET}")

    print(f"\n  Los resultados están en backtest/results/")
    print(f"  Para re-comparar: python scripts/compare_results.py --from-file {manifest_path.name}")
    print(f"  Abrí el dashboard para visualizarlos.\n")


if __name__ == "__main__":
    main()
